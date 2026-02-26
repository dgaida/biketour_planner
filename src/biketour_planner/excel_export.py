"""Excel export functionality for the Bike Tour Planner."""

import json
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .excel_hyperlinks import create_tourist_sights_hyperlinks


def extract_city_name(address: str) -> str:
    """Extracts only the city name from a full address string.

    Args:
        address: Full address, e.g., "Street 1, 21000 Split, Croatia"

    Returns:
        The city name, e.g., "Split".
    """
    if not address:
        return ""

    # Format: "Street, ZIP City, Country"
    # We want only "City"

    # Remove country at the end (after last comma)
    parts = address.split(",")

    # Determine the part that might contain the city
    if len(parts) >= 2:
        # Take the second to last part (should be "ZIP City")
        city_part = parts[-2].strip()
    else:
        # If no commas, take the whole string
        city_part = address.strip()

    # Remove ZIP (leading digits and spaces)
    city_match = re.search(r"^\d+\s+(.+)$", city_part)
    if city_match:
        return city_match.group(1).strip()

    # Fallback to cleaned city part
    return city_part


def create_accommodation_text(booking: dict, use_symbols: bool = False) -> str:
    """Creates a formatted text for accommodation including name, address, and amenities.

    Args:
        booking: Dictionary containing booking information.
        use_symbols: Whether to use symbols instead of text for amenities.

    Returns:
        Formatted multi-line text for the Excel cell.
    """
    text_parts = []

    # Hotel name
    if booking.get("hotel_name"):
        text_parts.append(booking["hotel_name"])

    # Address
    if booking.get("address"):
        text_parts.append(booking["address"])

    # Amenities
    amenities = []
    if booking.get("has_washing_machine"):
        amenities.append("Wasch")
    if booking.get("has_kitchen"):
        amenities.append("🍳" if use_symbols else "Küche")
    if booking.get("has_towels"):
        amenities.append("🧺" if use_symbols else "Handtücher")
    if booking.get("has_breakfast"):
        amenities.append("Früh")
    if booking.get("checkin_time"):
        amenities.append("Checkin: " + booking["checkin_time"])

    if amenities:
        text_parts.append(", ".join(amenities))

    return "\n".join(text_parts)


def export_bookings_to_excel(json_path: Path, template_path: Path, output_path: Path, start_row: int = 4) -> None:
    """Exports booking information to an Excel file based on a template.

    Automatically inserts empty lines for days without a booking (between the departure date
    of one booking and the arrival date of the next).

    Args:
        json_path: Path to the JSON file with bookings.
        template_path: Path to the Excel template file.
        output_path: Path for the output Excel file.
        start_row: Row to start inserting data (1-based).
    """
    # Load JSON
    with open(json_path, encoding="utf-8") as f:
        bookings = json.load(f)

    # Sort by arrival date
    bookings_sorted = sorted(bookings, key=lambda x: x.get("arrival_date", "9999-12-31"))

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active

    # Previous city for start-destination determination
    previous_city = None
    previous_departure_date = None

    # Current day counter
    day_counter = 1

    # Current Excel row
    current_row = start_row

    # Insert data
    for booking in bookings_sorted:
        # Check if empty rows for gap days need to be inserted
        if previous_departure_date and booking.get("arrival_date"):
            try:
                prev_departure = datetime.fromisoformat(previous_departure_date)
                current_arrival = datetime.fromisoformat(booking["arrival_date"])

                # Calculate difference in days
                days_between = (current_arrival - prev_departure).days

                # Insert empty rows (one row per gap day)
                if days_between > 0:
                    for day_offset in range(days_between):
                        # Column A: Day counter
                        ws[f"A{current_row}"] = day_counter

                        # Column B: Date of the gap day
                        intermediate_date = prev_departure + timedelta(days=day_offset)
                        ws[f"B{current_row}"] = intermediate_date
                        ws[f"B{current_row}"].number_format = "DDD, DD.MM.YYYY"

                        # Column C: Start city (previous city)
                        if previous_city:
                            ws[f"C{current_row}"] = previous_city

                        day_counter += 1
                        current_row += 1

            except ValueError:
                pass

        # Insert normal booking row
        row = current_row

        # Column A: Day counter
        ws[f"A{row}"] = day_counter

        # Column B: Date
        arrival_date = booking.get("arrival_date", "")
        if arrival_date:
            try:
                date_obj = datetime.fromisoformat(arrival_date)
                ws[f"B{row}"] = date_obj
                ws[f"B{row}"].number_format = "DDD, DD.MM.YYYY"
            except ValueError:
                ws[f"B{row}"] = arrival_date

        # Column C: Start city (previous city)
        if previous_city:
            ws[f"C{row}"] = previous_city

        # Column D: Destination city (current city)
        current_city = extract_city_name(booking.get("address", ""))
        ws[f"D{row}"] = current_city

        # Column E: distance in km
        ws[f"E{row}"] = booking.get("total_distance_km", "")

        # Column F: Accommodation with name, address, and amenities
        accommodation_text = create_accommodation_text(booking)
        ws[f"F{row}"] = accommodation_text
        ws[f"F{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        # Column G: Ascent / Max Elevation
        ws[f"G{row}"] = f"{booking.get('total_ascent_m', '')} / {booking.get('max_elevation_m', '')}"

        # Column H: Final track name
        ws[f"H{row}"] = booking.get("gpx_track_final", "")[:12]

        # Column I: Tourist Sights
        create_tourist_sights_hyperlinks(ws, row, booking.get("tourist_sights", None))

        # Column J: Price
        ws[f"J{row}"] = booking.get("total_price", "")

        # Column K: Cancellation info
        ws[f"K{row}"] = f"Stornierung bis: {booking.get('free_cancel_until', '')}"

        # Update variables for next iteration
        previous_city = current_city
        previous_departure_date = booking.get("departure_date")
        day_counter += 1
        current_row += 1

    # Save output
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
