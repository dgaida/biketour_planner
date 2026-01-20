"""Excel-Info-Reader für zusätzliche Tagesinformationen.

Dieses Modul liest eine Excel-Datei mit zusätzlichen Infos pro Tag ein
und bereitet diese für den PDF-Export auf.
"""

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook


def read_daily_info_from_excel(excel_path: Path, start_row: int = 2) -> dict[str, list[str]]:
    """Liest zusätzliche Tagesinformationen aus Excel-Datei.

    Die Funktion liest Spalte B (Datum) und Spalte C (Infos) aus der Excel-Datei.
    Infos in Spalte C sind durch Semikolon getrennt. Links (beginnend mit "http")
    werden als HTML-Links formatiert.

    Args:
        excel_path: Pfad zur Excel-Datei mit Tagesinformationen.
        start_row: Zeile ab der die Daten gelesen werden (Default: 2).

    Returns:
        Dictionary mit ISO-Datum als Key und Liste von Infos/Links als Value.

        Struktur:
            {
                "2026-05-15": [
                    "Markt besuchen",
                    '<a href="https://example.com" color="blue"><u>Restaurant-Tipp</u></a>',
                    "Früh losfahren"
                ],
                "2026-05-16": [...],
                ...
            }

    Example:
        >>> infos = read_daily_info_from_excel(Path("Reiseplanung_Fahrrad.xlsx"))
        >>> print(infos["2026-05-15"])
        ['Markt besuchen', '<a href="..." color="blue"><u>Restaurant-Tipp</u></a>']
    """
    if not excel_path.exists():
        print(f"⚠️  Excel-Datei nicht gefunden: {excel_path}")
        return {}

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        daily_info = {}

        for row in range(start_row, ws.max_row + 1):
            # Lese Datum aus Spalte B
            date_cell = ws[f"B{row}"]
            info_cell = ws[f"C{row}"]

            # Überspringe leere Zeilen
            if not date_cell.value:
                continue

            # Konvertiere Datum zu ISO-Format (YYYY-MM-DD)
            try:
                if isinstance(date_cell.value, datetime):
                    date_str = date_cell.value.strftime("%Y-%m-%d")
                else:
                    # Versuche String zu parsen
                    date_obj = datetime.strptime(str(date_cell.value), "%Y-%m-%d")
                    date_str = date_obj.strftime("%Y-%m-%d")
            except (ValueError, AttributeError):
                print(f"⚠️  Konnte Datum in Zeile {row} nicht parsen: {date_cell.value}")
                continue

            # Lese Infos aus Spalte C
            if not info_cell.value:
                continue

            info_text = str(info_cell.value).strip()

            # Splitte bei Semikolon
            info_items = [item.strip() for item in info_text.split(";") if item.strip()]

            # Konvertiere Links zu HTML
            formatted_items = []
            for item in info_items:
                if item.startswith("http"):
                    # Extrahiere Linktext (nach letztem Slash oder verwende URL)
                    link_text = item.split("/")[-1] if "/" in item else item
                    # Begrenze Linktext auf 50 Zeichen
                    if len(link_text) > 50:
                        link_text = link_text[:47] + "..."

                    html_link = f'<a href="{item}" color="blue"><u>{link_text}</u></a>'
                    formatted_items.append(html_link)
                else:
                    formatted_items.append(item)

            # Füge zum Dictionary hinzu
            if formatted_items:
                daily_info[date_str] = formatted_items

        print(f"✅ {len(daily_info)} Tage mit Zusatzinfos geladen")
        return daily_info

    except Exception as e:
        print(f"❌ Fehler beim Lesen der Excel-Datei: {e}")
        return {}


if __name__ == "__main__":
    # Test
    test_path = Path("../2026_Kroatien/booking/Reiseplanung_Fahrrad.xlsx")
    infos = read_daily_info_from_excel(test_path)

    for date, items in infos.items():
        print(f"\n{date}:")
        for item in items:
            print(f"  - {item}")
