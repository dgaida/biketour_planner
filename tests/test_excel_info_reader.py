"""Unit-Tests für excel_info_reader.py.

Testet das Lesen von zusätzlichen Tagesinformationen aus Excel-Dateien inklusive:
- Lesen von Datum und Info-Spalten
- Verarbeitung von Semikolon-getrennten Einträgen
- Konvertierung von URLs zu HTML-Links
- Fehlerbehandlung bei ungültigen Daten
"""

from datetime import datetime

import pytest
from openpyxl import Workbook

from biketour_planner.excel_info_reader import read_daily_info_from_excel

# ============================================================================
# Test-Fixtures
# ============================================================================


@pytest.fixture
def simple_excel_file(tmp_path):
    """Erstellt eine einfache Test-Excel-Datei mit Tagesinformationen."""
    wb = Workbook()
    ws = wb.active

    # Header
    ws["A1"] = "Tag"
    ws["B1"] = "Datum"
    ws["C1"] = "Infos"

    # Daten
    ws["A2"] = 1
    ws["B2"] = datetime(2026, 5, 15)
    ws["C2"] = "Markt besuchen; Früh losfahren"

    ws["A3"] = 2
    ws["B3"] = datetime(2026, 5, 16)
    ws["C3"] = "Restaurant-Tipp; https://example.com/restaurant"

    ws["A4"] = 3
    ws["B4"] = datetime(2026, 5, 17)
    ws["C4"] = "Ruhetag"

    excel_file = tmp_path / "test_info.xlsx"
    wb.save(excel_file)
    return excel_file


@pytest.fixture
def excel_with_links(tmp_path):
    """Erstellt Excel-Datei mit mehreren URLs."""
    wb = Workbook()
    ws = wb.active

    ws["B2"] = datetime(2026, 5, 15)
    ws["C2"] = "https://example.com/info1; https://example.com/info2"

    excel_file = tmp_path / "links.xlsx"
    wb.save(excel_file)
    return excel_file


@pytest.fixture
def excel_with_empty_cells(tmp_path):
    """Erstellt Excel-Datei mit leeren Zellen."""
    wb = Workbook()
    ws = wb.active

    ws["B2"] = datetime(2026, 5, 15)
    ws["C2"] = "Info 1"

    ws["B3"] = datetime(2026, 5, 16)
    # C3 ist leer

    ws["B4"] = None  # Leeres Datum
    ws["C4"] = "Info 3"

    excel_file = tmp_path / "empty_cells.xlsx"
    wb.save(excel_file)
    return excel_file


# ============================================================================
# Test read_daily_info_from_excel
# ============================================================================


class TestReadDailyInfoFromExcel:
    """Tests für die read_daily_info_from_excel Funktion."""

    def test_read_basic_info(self, simple_excel_file):
        """Testet grundlegendes Lesen von Tagesinformationen."""
        result = read_daily_info_from_excel(simple_excel_file)

        assert "2026-05-15" in result
        assert "2026-05-16" in result
        assert "2026-05-17" in result
        assert len(result) == 3

    def test_read_splits_by_semicolon(self, simple_excel_file):
        """Testet dass Einträge bei Semikolon getrennt werden."""
        result = read_daily_info_from_excel(simple_excel_file)

        assert len(result["2026-05-15"]) == 2
        assert "Markt besuchen" in result["2026-05-15"]
        assert "Früh losfahren" in result["2026-05-15"]

    def test_read_converts_urls_to_links(self, simple_excel_file):
        """Testet dass URLs zu HTML-Links konvertiert werden."""
        result = read_daily_info_from_excel(simple_excel_file)

        # Zweiter Eintrag sollte Link sein
        link_entry = result["2026-05-16"][1]
        assert "<a href=" in link_entry
        assert "https://example.com/restaurant" in link_entry
        assert "<u>" in link_entry
        assert 'color="blue"' in link_entry

    def test_read_preserves_non_link_text(self, simple_excel_file):
        """Testet dass normaler Text erhalten bleibt."""
        result = read_daily_info_from_excel(simple_excel_file)

        assert "Markt besuchen" == result["2026-05-15"][0]
        assert "Restaurant-Tipp" == result["2026-05-16"][0]
        assert "Ruhetag" == result["2026-05-17"][0]

    def test_read_multiple_links(self, excel_with_links):
        """Testet Verarbeitung mehrerer URLs."""
        result = read_daily_info_from_excel(excel_with_links)

        infos = result["2026-05-15"]
        assert len(infos) == 2
        assert all("<a href=" in info for info in infos)

    def test_read_date_formats(self, tmp_path):
        """Testet verschiedene Datumsformate."""
        wb = Workbook()
        ws = wb.active

        # Verschiedene Datumsformate
        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "Info 1"

        ws["B3"] = "2026-05-16"  # String-Datum
        ws["C3"] = "Info 2"

        excel_file = tmp_path / "dates.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        assert "2026-05-15" in result
        # String-Datum sollte auch funktionieren
        assert "2026-05-16" in result

    def test_read_skips_empty_info_cells(self, excel_with_empty_cells):
        """Testet dass leere Info-Zellen übersprungen werden."""
        result = read_daily_info_from_excel(excel_with_empty_cells)

        # Nur Zeilen mit Info sollten vorhanden sein
        assert "2026-05-15" in result
        assert "2026-05-16" not in result  # Leere Info-Zelle

    def test_read_skips_empty_date_cells(self, excel_with_empty_cells):
        """Testet dass leere Datumszellen übersprungen werden."""
        result = read_daily_info_from_excel(excel_with_empty_cells)

        # Zeile mit leerem Datum sollte nicht enthalten sein
        assert len(result) == 1
        assert "2026-05-15" in result

    def test_read_trims_whitespace(self, tmp_path):
        """Testet dass Whitespace entfernt wird."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "  Info mit Spaces  ;  Zweiter Eintrag  "

        excel_file = tmp_path / "whitespace.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        assert "Info mit Spaces" in result["2026-05-15"]
        assert "Zweiter Eintrag" in result["2026-05-15"]
        # Keine führenden/abschließenden Spaces
        assert not any(item.startswith(" ") for item in result["2026-05-15"])

    def test_read_handles_long_urls(self, tmp_path):
        """Testet Behandlung sehr langer URLs."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        long_url = "https://example.com/" + "a" * 100
        ws["C2"] = long_url

        excel_file = tmp_path / "long_url.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        link = result["2026-05-15"][0]
        assert "<a href=" in link
        # Link-Text sollte auf 50 Zeichen begrenzt sein
        assert "..." in link

    def test_read_custom_start_row(self, tmp_path):
        """Testet benutzerdefinierte Start-Zeile."""
        wb = Workbook()
        ws = wb.active

        # Header in Zeile 3
        ws["B3"] = "Datum"
        ws["C3"] = "Info"

        # Daten ab Zeile 4
        ws["B4"] = datetime(2026, 5, 15)
        ws["C4"] = "Info ab Zeile 4"

        excel_file = tmp_path / "custom_start.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file, start_row=4)

        assert "2026-05-15" in result
        assert "Info ab Zeile 4" in result["2026-05-15"]

    def test_read_file_not_found(self, tmp_path):
        """Testet Verhalten bei nicht existierender Datei."""
        non_existent = tmp_path / "does_not_exist.xlsx"

        result = read_daily_info_from_excel(non_existent)

        assert result == {}

    def test_read_invalid_excel_file(self, tmp_path):
        """Testet Verhalten bei ungültiger Excel-Datei."""
        invalid_file = tmp_path / "invalid.xlsx"
        invalid_file.write_text("nicht excel", encoding="utf-8")

        result = read_daily_info_from_excel(invalid_file)

        assert result == {}

    def test_read_empty_excel_file(self, tmp_path):
        """Testet Verhalten bei leerer Excel-Datei."""
        wb = Workbook()
        ws = wb.active
        print(ws)

        excel_file = tmp_path / "empty.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        assert result == {}

    def test_read_invalid_date_format(self, tmp_path):
        """Testet Verhalten bei ungültigem Datumsformat."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = "ungültiges datum"
        ws["C2"] = "Info"

        excel_file = tmp_path / "invalid_date.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        # Zeile mit ungültigem Datum sollte übersprungen werden
        assert result == {}

    def test_read_semicolon_in_text(self, tmp_path):
        """Testet Behandlung von Semikolon in normalem Text."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "Text mit ; im Inhalt; Zweiter Teil"

        excel_file = tmp_path / "semicolon.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        # Sollte bei jedem Semikolon trennen
        assert len(result["2026-05-15"]) == 3

    def test_read_only_whitespace_entries(self, tmp_path):
        """Testet dass nur Whitespace-Einträge übersprungen werden."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "Info 1;   ;Info 2"

        excel_file = tmp_path / "whitespace_only.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        # Nur nicht-leere Einträge
        assert len(result["2026-05-15"]) == 2
        assert "Info 1" in result["2026-05-15"]
        assert "Info 2" in result["2026-05-15"]

    def test_read_url_detection(self, tmp_path):
        """Testet URL-Erkennung in verschiedenen Formaten."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "http://example.com; https://secure.com; ftp://files.com"

        excel_file = tmp_path / "urls.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        infos = result["2026-05-15"]
        # Alle URLs sollten zu Links konvertiert werden
        assert all("<a href=" in info for info in infos)
        assert len(infos) == 3

    def test_read_mixed_links_and_text(self, tmp_path):
        """Testet gemischte URLs und normalen Text."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "Normaler Text; https://example.com; Mehr Text"

        excel_file = tmp_path / "mixed.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        infos = result["2026-05-15"]
        assert len(infos) == 3
        assert "Normaler Text" == infos[0]
        assert "<a href=" in infos[1]
        assert "Mehr Text" == infos[2]

    def test_read_date_iso_format(self, tmp_path):
        """Testet ISO-Datumsformat im Output."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 12, 31)
        ws["C2"] = "Silvester"

        excel_file = tmp_path / "iso_date.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        # Datum sollte im ISO-Format sein
        assert "2026-12-31" in result

    def test_read_preserves_info_order(self, tmp_path):
        """Testet dass Reihenfolge der Infos erhalten bleibt."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "Erster; Zweiter; Dritter"

        excel_file = tmp_path / "order.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        assert result["2026-05-15"][0] == "Erster"
        assert result["2026-05-15"][1] == "Zweiter"
        assert result["2026-05-15"][2] == "Dritter"

    def test_read_unicode_characters(self, tmp_path):
        """Testet Umgang mit Unicode-Zeichen."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "Café besuchen; Müller-Straße; 北京"

        excel_file = tmp_path / "unicode.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        assert "Café besuchen" in result["2026-05-15"]
        assert "Müller-Straße" in result["2026-05-15"]
        assert "北京" in result["2026-05-15"]

    def test_read_special_characters_in_url(self, tmp_path):
        """Testet URLs mit Sonderzeichen."""
        wb = Workbook()
        ws = wb.active

        ws["B2"] = datetime(2026, 5, 15)
        ws["C2"] = "https://example.com/path?param=value&other=123"

        excel_file = tmp_path / "url_special.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        link = result["2026-05-15"][0]
        assert '<a href="https://example.com/path?param=value&other=123"' in link

    def test_read_multiple_dates(self, tmp_path):
        """Testet Lesen mehrerer Tage."""
        wb = Workbook()
        ws = wb.active

        for i in range(1, 11):
            ws[f"B{i + 1}"] = datetime(2026, 5, i)
            ws[f"C{i + 1}"] = f"Info für Tag {i}"

        excel_file = tmp_path / "multiple.xlsx"
        wb.save(excel_file)

        result = read_daily_info_from_excel(excel_file)

        assert len(result) == 10
        for i in range(1, 11):
            date_key = f"2026-05-{i:02d}"
            assert date_key in result

    def test_read_returns_dict(self, simple_excel_file):
        """Testet dass Dictionary zurückgegeben wird."""
        result = read_daily_info_from_excel(simple_excel_file)

        assert isinstance(result, dict)
        assert all(isinstance(k, str) for k in result.keys())
        assert all(isinstance(v, list) for v in result.values())


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
